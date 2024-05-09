from tkinter import filedialog
import pandas as pd
from datetime import datetime
from tkinter import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#Nesse arquivo estão as funções que auxiliam na criação da planilha excel
#com os pedidos de suprimento.

def recuperarHoraAtual():
    data_hora_atual = datetime.now()
    formato = "%Y-%m-%d_%H-%M-%S"
    data_hora_formatada = data_hora_atual.strftime(formato)
    return data_hora_formatada

def formatarTabela(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    ws['A1'] = 'ID'
    ws['B1'] = 'Nome do produto'
    ws['C1'] = 'Classificação'
    ws['D1'] = 'Linha'
    ws['E1'] = 'Estoque'
    ws['F1'] = 'Un. Estoque'
    ws['G1'] = 'Qtd. Produção'
    ws['H1'] = 'Unidade'
    
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="FDDA0D", end_color="FDDA0D", fill_type="solid")
    for cell in ws['E'][1:]:
        cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
    for cell in ws['F'][1:]:
        cell.fill = PatternFill(start_color="5D8AA8", end_color="5D8AA8", fill_type="solid")
    for cell in ws['G'][1:]:
        cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
    for cell in ws['H'][1:]:
        cell.fill = PatternFill(start_color="6b8e23", end_color="6b8e23", fill_type="solid")
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    wb.save(caminho)


def gerarArquivoExcel(tipoArquivo, listaProdutos):
    root = Tk()
    root.withdraw()

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                               filetypes=[("Arquivos Excel", "*.xlsx")],
                                               title="Salvar Arquivo Excel",
                                               initialfile=f"{tipoArquivo}--{recuperarHoraAtual()}")

    if not file_path:
        print("Operação cancelada pelo usuário.")
        return

    if not file_path.endswith(".xlsx"):
        file_path += ".xlsx"
        
    formatoTabela = pd.DataFrame(listaProdutos)
    formatoTabela = formatoTabela[['idProdutoComposicao', 'nomeProdutoComposicao', 'classificacao', 'linha', 'estoque', 'unidadeEstoque', 'totalProducao', 'unidade']]
    formatoTabela.to_excel(file_path, index=False)
    formatarTabela(file_path)
    print(f"Arquivo salvo em: {file_path}")

somaProdutosEventos = []
